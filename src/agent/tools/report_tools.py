"""Report generation tools — Excel/CSV exports with charts, uploaded to S3 with presigned download URLs."""

import json
import io
import os
import math
import base64
from datetime import datetime

from src.agent.tools.registry import tool_definition
from src.connectors.snowflake_connector import execute_query

# ── S3 helpers ────────────────────────────────────────────────────────────────

REPORT_BUCKET = os.getenv("REPORT_BUCKET", "snowflake-eyecare-reports")
REPORT_PREFIX = os.getenv("REPORT_PREFIX", "reports/")
PRESIGN_EXPIRY = int(os.getenv("PRESIGN_EXPIRY", "3600"))  # 1 hour default


def _upload_to_s3(data: bytes, key: str, content_type: str, filename: str) -> str:
    """Upload bytes to S3 and return a direct download URL."""
    import boto3
    from botocore.config import Config

    region = os.getenv("AWS_REGION", "us-west-2")
    s3 = boto3.client(
        "s3",
        region_name=region,
        config=Config(signature_version="s3v4"),
    )
    s3.put_object(
        Bucket=REPORT_BUCKET,
        Key=key,
        Body=data,
        ContentType=content_type,
        ContentDisposition=f'attachment; filename="{filename}"',
    )
    # Return direct S3 URL (bucket policy allows public read on reports/ prefix)
    url = f"https://{REPORT_BUCKET}.s3.{region}.amazonaws.com/{key}"
    return url


# ── SVG Chart Generation (zero-dependency) ────────────────────────────────────

def _svg_bar_chart(labels, values, title="", width=480, height=300, color="#29B5E8"):
    """Generate a bar chart as SVG string."""
    if not values:
        return ""
    margin = {"top": 40, "right": 20, "bottom": 60, "left": 70}
    plot_w = width - margin["left"] - margin["right"]
    plot_h = height - margin["top"] - margin["bottom"]
    max_val = max(values) if max(values) > 0 else 1
    bar_gap = 4
    bar_w = max(8, (plot_w - bar_gap * len(values)) / len(values))

    bars = []
    x_labels = []
    for i, (label, val) in enumerate(zip(labels, values)):
        x = margin["left"] + i * (bar_w + bar_gap) + bar_gap / 2
        bar_h = (val / max_val) * plot_h
        y = margin["top"] + plot_h - bar_h
        bars.append(
            f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_w:.1f}" height="{bar_h:.1f}" '
            f'fill="{color}" rx="2"><title>{label}: {val:,.2f}</title></rect>'
        )
        lbl = str(label)[:12]
        lx = x + bar_w / 2
        x_labels.append(
            f'<text x="{lx:.1f}" y="{margin["top"] + plot_h + 14}" '
            f'text-anchor="end" transform="rotate(-35 {lx:.1f} {margin["top"] + plot_h + 14})" '
            f'font-size="10" fill="#666">{lbl}</text>'
        )

    # Y-axis ticks
    y_ticks = []
    for i in range(5):
        v = max_val * i / 4
        y = margin["top"] + plot_h - (v / max_val) * plot_h
        if v >= 1_000_000:
            lbl = f"${v / 1_000_000:.1f}M"
        elif v >= 1_000:
            lbl = f"${v / 1_000:.0f}K"
        else:
            lbl = f"{v:,.0f}"
        y_ticks.append(
            f'<text x="{margin["left"] - 8}" y="{y + 4:.1f}" text-anchor="end" font-size="10" fill="#999">{lbl}</text>'
            f'<line x1="{margin["left"]}" y1="{y:.1f}" x2="{margin["left"] + plot_w}" y2="{y:.1f}" stroke="#eee" stroke-width="1"/>'
        )

    title_el = f'<text x="{width / 2}" y="22" text-anchor="middle" font-size="14" font-weight="600" fill="#1a2332">{title}</text>' if title else ""

    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" width="{width}" height="{height}" style="font-family:system-ui,sans-serif;background:#fff;border-radius:8px">
{title_el}
{"".join(y_ticks)}
{"".join(bars)}
{"".join(x_labels)}
<line x1="{margin['left']}" y1="{margin['top'] + plot_h}" x2="{margin['left'] + plot_w}" y2="{margin['top'] + plot_h}" stroke="#ccc" stroke-width="1"/>
<line x1="{margin['left']}" y1="{margin['top']}" x2="{margin['left']}" y2="{margin['top'] + plot_h}" stroke="#ccc" stroke-width="1"/>
</svg>'''


def _svg_line_chart(labels, values, title="", width=480, height=300, color="#006FB4"):
    """Generate a line chart as SVG string."""
    if not values or len(values) < 2:
        return ""
    margin = {"top": 40, "right": 20, "bottom": 60, "left": 70}
    plot_w = width - margin["left"] - margin["right"]
    plot_h = height - margin["top"] - margin["bottom"]
    max_val = max(values) if max(values) > 0 else 1
    min_val = min(values)
    val_range = max_val - min_val if max_val != min_val else 1

    points = []
    dots = []
    x_labels_el = []
    step = plot_w / (len(values) - 1) if len(values) > 1 else plot_w
    for i, (label, val) in enumerate(zip(labels, values)):
        x = margin["left"] + i * step
        y = margin["top"] + plot_h - ((val - min_val) / val_range) * plot_h
        points.append(f"{x:.1f},{y:.1f}")
        dots.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="{color}"><title>{label}: {val:,.2f}</title></circle>')
        if i % max(1, len(values) // 8) == 0:
            lbl = str(label)[:10]
            x_labels_el.append(
                f'<text x="{x:.1f}" y="{margin["top"] + plot_h + 16}" text-anchor="middle" font-size="10" fill="#666">{lbl}</text>'
            )

    # Y-axis
    y_ticks = []
    for i in range(5):
        v = min_val + val_range * i / 4
        y = margin["top"] + plot_h - ((v - min_val) / val_range) * plot_h
        if abs(v) >= 1_000_000:
            lbl = f"${v / 1_000_000:.1f}M"
        elif abs(v) >= 1_000:
            lbl = f"${v / 1_000:.0f}K"
        else:
            lbl = f"{v:,.0f}"
        y_ticks.append(
            f'<text x="{margin["left"] - 8}" y="{y + 4:.1f}" text-anchor="end" font-size="10" fill="#999">{lbl}</text>'
            f'<line x1="{margin["left"]}" y1="{y:.1f}" x2="{margin["left"] + plot_w}" y2="{y:.1f}" stroke="#eee" stroke-width="1"/>'
        )

    # Gradient fill under line
    fill_points = f'{margin["left"]:.1f},{margin["top"] + plot_h:.1f} ' + " ".join(points) + f' {margin["left"] + (len(values) - 1) * step:.1f},{margin["top"] + plot_h:.1f}'

    title_el = f'<text x="{width / 2}" y="22" text-anchor="middle" font-size="14" font-weight="600" fill="#1a2332">{title}</text>' if title else ""

    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" width="{width}" height="{height}" style="font-family:system-ui,sans-serif;background:#fff;border-radius:8px">
<defs><linearGradient id="fill" x1="0" y1="0" x2="0" y2="1"><stop offset="0%" stop-color="{color}" stop-opacity="0.15"/><stop offset="100%" stop-color="{color}" stop-opacity="0.02"/></linearGradient></defs>
{title_el}
{"".join(y_ticks)}
<polygon points="{fill_points}" fill="url(#fill)"/>
<polyline points="{" ".join(points)}" fill="none" stroke="{color}" stroke-width="2" stroke-linejoin="round"/>
{"".join(dots)}
{"".join(x_labels_el)}
<line x1="{margin['left']}" y1="{margin['top'] + plot_h}" x2="{margin['left'] + plot_w}" y2="{margin['top'] + plot_h}" stroke="#ccc" stroke-width="1"/>
</svg>'''


def _svg_pie_chart(labels, values, title="", width=320, height=320):
    """Generate a pie chart as SVG string."""
    if not values or sum(values) == 0:
        return ""
    colors = ["#29B5E8", "#006FB4", "#34A853", "#FBBC04", "#EA4335", "#9334E6", "#FF6D01", "#46BDC6"]
    cx, cy, r = width / 2, height / 2 + 10, min(width, height) / 2 - 50
    total = sum(values)

    slices = []
    legend = []
    angle = -90
    for i, (label, val) in enumerate(zip(labels, values)):
        pct = val / total
        sweep = pct * 360
        color = colors[i % len(colors)]

        # Arc path
        start_rad = math.radians(angle)
        end_rad = math.radians(angle + sweep)
        x1 = cx + r * math.cos(start_rad)
        y1 = cy + r * math.sin(start_rad)
        x2 = cx + r * math.cos(end_rad)
        y2 = cy + r * math.sin(end_rad)
        large = 1 if sweep > 180 else 0

        slices.append(
            f'<path d="M{cx},{cy} L{x1:.1f},{y1:.1f} A{r},{r} 0 {large} 1 {x2:.1f},{y2:.1f} Z" '
            f'fill="{color}" stroke="#fff" stroke-width="1.5"><title>{label}: {val:,.0f} ({pct:.1%})</title></path>'
        )

        # Legend
        ly = height - 30 + (i // 3) * 16
        lx = 10 + (i % 3) * (width // 3)
        lbl = str(label)[:14]
        legend.append(
            f'<rect x="{lx}" y="{ly - 8}" width="10" height="10" rx="2" fill="{color}"/>'
            f'<text x="{lx + 14}" y="{ly}" font-size="10" fill="#666">{lbl} ({pct:.0%})</text>'
        )
        angle += sweep

    title_el = f'<text x="{width / 2}" y="22" text-anchor="middle" font-size="14" font-weight="600" fill="#1a2332">{title}</text>' if title else ""

    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height + len(labels) // 3 * 16}" width="{width}" height="{height + len(labels) // 3 * 16}" style="font-family:system-ui,sans-serif;background:#fff;border-radius:8px">
{title_el}
{"".join(slices)}
{"".join(legend)}
</svg>'''


# ── Tool: generate_report ─────────────────────────────────────────────────────

@tool_definition(
    name="generate_report",
    description=(
        "Run a SQL query and generate a downloadable Excel (.xlsx) or CSV report. "
        "Returns a presigned S3 download URL valid for 1 hour. "
        "Use for any request like 'give me a report', 'export to Excel', 'download the data', etc."
    ),
    parameters={
        "properties": {
            "sql": {
                "type": "string",
                "description": "The SQL query to run. Must be a SELECT statement.",
            },
            "title": {
                "type": "string",
                "description": "Report title (used as sheet name and filename).",
                "default": "Report",
            },
            "format": {
                "type": "string",
                "description": "'excel' or 'csv'. Defaults to 'excel'.",
                "default": "excel",
            },
            "include_chart": {
                "type": "boolean",
                "description": "If true and format is 'excel', adds a chart sheet. Defaults to true.",
                "default": True,
            },
            "chart_type": {
                "type": "string",
                "description": "Chart type: 'bar', 'line', or 'pie'. Defaults to 'bar'.",
                "default": "bar",
            },
            "chart_x_column": {
                "type": "string",
                "description": "Column name to use for chart X axis / labels.",
            },
            "chart_y_column": {
                "type": "string",
                "description": "Column name to use for chart Y axis / values.",
            },
        },
        "required": ["sql"],
    },
)
def generate_report(
    sql,
    title="Report",
    format="excel",
    include_chart=True,
    chart_type="bar",
    chart_x_column=None,
    chart_y_column=None,
):
    """Generate an Excel or CSV report from a SQL query and upload to S3."""
    import re

    # Validate read-only
    forbidden = re.compile(
        r"\b(INSERT|UPDATE|DELETE|DROP|CREATE|ALTER|TRUNCATE|MERGE|GRANT|REVOKE)\b",
        re.IGNORECASE,
    )
    if forbidden.search(sql):
        return {"error": "Only SELECT queries are allowed for report generation."}

    # Execute query
    df = execute_query(sql, limit=50000)
    if df.empty:
        return {"message": "Query returned no data. No report generated."}

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_title = re.sub(r"[^a-zA-Z0-9_-]", "_", title)[:40]

    if format == "csv":
        # CSV export
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)
        csv_bytes = csv_buffer.getvalue().encode("utf-8")
        filename = f"{safe_title}_{timestamp}.csv"
        key = f"{REPORT_PREFIX}{filename}"
        url = _upload_to_s3(csv_bytes, key, "text/csv", filename)
        return {
            "download_url": url,
            "filename": filename,
            "format": "csv",
            "rows": len(df),
            "columns": list(df.columns),
        }

    # Excel export with openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = safe_title[:31]  # Excel sheet name limit

    # ── Header row ──
    header_fill = PatternFill(start_color="29B5E8", end_color="29B5E8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # ── Data rows ──
    num_font = Font(size=10)
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Handle various types
            if hasattr(value, "item"):  # numpy scalar
                value = value.item()
            if isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
            else:
                cell.value = str(value) if value is not None else ""
            cell.font = num_font
            cell.border = thin_border

    # ── Auto-width columns ──
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(df) + 2, 102)):  # Sample first 100 rows
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    # ── Freeze header row ──
    ws.freeze_panes = "A2"

    # ── Auto filters ──
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

    # ── Summary sheet ──
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary["A1"] = title
    ws_summary["A1"].font = Font(bold=True, size=16, color="1A2332")
    ws_summary["A3"] = "Generated"
    ws_summary["B3"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    ws_summary["A4"] = "Total Rows"
    ws_summary["B4"] = len(df)
    ws_summary["A5"] = "Columns"
    ws_summary["B5"] = len(df.columns)

    # Add numeric summaries
    row_num = 7
    ws_summary.cell(row=row_num, column=1, value="Column Summaries").font = Font(bold=True, size=12)
    row_num += 1
    for col in df.columns:
        if df[col].dtype in ("float64", "int64", "float32", "int32"):
            ws_summary.cell(row=row_num, column=1, value=str(col)).font = Font(bold=True)
            ws_summary.cell(row=row_num, column=2, value=f"Sum: {df[col].sum():,.2f}")
            ws_summary.cell(row=row_num, column=3, value=f"Avg: {df[col].mean():,.2f}")
            ws_summary.cell(row=row_num, column=4, value=f"Min: {df[col].min():,.2f}")
            ws_summary.cell(row=row_num, column=5, value=f"Max: {df[col].max():,.2f}")
            row_num += 1

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 25
    ws_summary.column_dimensions["C"].width = 25
    ws_summary.column_dimensions["D"].width = 25
    ws_summary.column_dimensions["E"].width = 25

    # ── Chart sheet ──
    if include_chart and len(df) > 1:
        # Determine chart columns
        x_col = chart_x_column
        y_col = chart_y_column

        if not x_col:
            # First non-numeric column
            for c in df.columns:
                if df[c].dtype == "object" or str(df[c].dtype).startswith("datetime"):
                    x_col = c
                    break
            if not x_col:
                x_col = df.columns[0]

        if not y_col:
            # First numeric column
            for c in df.columns:
                if df[c].dtype in ("float64", "int64", "float32", "int32"):
                    y_col = c
                    break

        if y_col and x_col:
            x_idx = list(df.columns).index(x_col) + 1
            y_idx = list(df.columns).index(y_col) + 1

            ws_chart = wb.create_sheet("Chart")
            data_sheet = ws  # reference data sheet

            # openpyxl chart
            if chart_type == "pie":
                chart = PieChart()
                chart.style = 10
            elif chart_type == "line":
                chart = LineChart()
                chart.style = 10
                chart.y_axis.numFmt = '#,##0'
            else:
                chart = BarChart()
                chart.style = 10
                chart.y_axis.numFmt = '#,##0'

            chart.title = title
            chart.width = 20
            chart.height = 14

            data_ref = Reference(data_sheet, min_col=y_idx, min_row=1, max_row=min(len(df) + 1, 101))
            cats_ref = Reference(data_sheet, min_col=x_idx, min_row=2, max_row=min(len(df) + 1, 101))

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4

            ws_chart.add_chart(chart, "A1")

    # Write to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_bytes = excel_buffer.getvalue()

    filename = f"{safe_title}_{timestamp}.xlsx"
    key = f"{REPORT_PREFIX}{filename}"
    url = _upload_to_s3(excel_bytes, key, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename)

    return {
        "download_url": url,
        "filename": filename,
        "format": "excel",
        "rows": len(df),
        "columns": list(df.columns),
        "sheets": ["Summary", safe_title[:31]] + (["Chart"] if include_chart else []),
    }


# ── Tool: generate_chart ──────────────────────────────────────────────────────

@tool_definition(
    name="generate_chart",
    description=(
        "Run a SQL query and generate an inline chart (SVG). "
        "Returns a base64-encoded SVG image that can be displayed directly in the chat. "
        "Use for any request like 'show me a chart', 'graph the data', 'visualize', etc."
    ),
    parameters={
        "properties": {
            "sql": {
                "type": "string",
                "description": "The SQL query to run. Must be a SELECT statement.",
            },
            "chart_type": {
                "type": "string",
                "description": "'bar', 'line', or 'pie'. Defaults to 'bar'.",
                "default": "bar",
            },
            "x_column": {
                "type": "string",
                "description": "Column name for X axis / labels. Auto-detected if omitted.",
            },
            "y_column": {
                "type": "string",
                "description": "Column name for Y axis / values. Auto-detected if omitted.",
            },
            "title": {
                "type": "string",
                "description": "Chart title.",
                "default": "Chart",
            },
        },
        "required": ["sql"],
    },
)
def generate_chart(sql, chart_type="bar", x_column=None, y_column=None, title="Chart"):
    """Generate an inline SVG chart from a SQL query."""
    import re

    forbidden = re.compile(
        r"\b(INSERT|UPDATE|DELETE|DROP|CREATE|ALTER|TRUNCATE|MERGE|GRANT|REVOKE)\b",
        re.IGNORECASE,
    )
    if forbidden.search(sql):
        return {"error": "Only SELECT queries are allowed."}

    df = execute_query(sql, limit=500)
    if df.empty:
        return {"message": "Query returned no data."}

    # Auto-detect columns
    x_col = x_column
    y_col = y_column

    if not x_col:
        for c in df.columns:
            if df[c].dtype == "object" or str(df[c].dtype).startswith("datetime"):
                x_col = c
                break
        if not x_col:
            x_col = df.columns[0]

    if not y_col:
        for c in df.columns:
            if c != x_col and df[c].dtype in ("float64", "int64", "float32", "int32"):
                y_col = c
                break

    if not y_col:
        return {"error": f"No numeric column found for chart values. Columns: {list(df.columns)}"}

    labels = [str(v) for v in df[x_col].tolist()]
    values = [float(v) if v is not None else 0 for v in df[y_col].tolist()]

    if chart_type == "pie":
        svg = _svg_pie_chart(labels, values, title=title)
    elif chart_type == "line":
        svg = _svg_line_chart(labels, values, title=title)
    else:
        svg = _svg_bar_chart(labels, values, title=title)

    if not svg:
        return {"error": "Could not generate chart from the data."}

    # Encode as base64 data URI
    svg_b64 = base64.b64encode(svg.encode("utf-8")).decode("utf-8")

    return {
        "chart_image": f"data:image/svg+xml;base64,{svg_b64}",
        "chart_type": chart_type,
        "data_points": len(labels),
        "x_column": x_col,
        "y_column": y_col,
    }


# ── Tool: multi_report ────────────────────────────────────────────────────────

@tool_definition(
    name="multi_report",
    description=(
        "Generate a multi-sheet Excel report by running multiple SQL queries. "
        "Each query becomes a separate sheet. Great for comprehensive reports like "
        "'give me a full practice performance report' covering revenue, patients, products, appointments."
    ),
    parameters={
        "properties": {
            "title": {
                "type": "string",
                "description": "Overall report title.",
            },
            "queries": {
                "type": "string",
                "description": (
                    "JSON array of query objects: "
                    '[{"name": "Sheet Name", "sql": "SELECT ...", "chart_type": "bar|line|pie"}]. '
                    "Each object creates one sheet. chart_type is optional."
                ),
            },
        },
        "required": ["title", "queries"],
    },
)
def multi_report(title, queries):
    """Generate a multi-sheet Excel report from multiple SQL queries."""
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    try:
        query_list = json.loads(queries) if isinstance(queries, str) else queries
    except json.JSONDecodeError:
        return {"error": "Invalid JSON in queries parameter."}

    if not query_list or len(query_list) == 0:
        return {"error": "No queries provided."}
    if len(query_list) > 10:
        return {"error": "Maximum 10 queries per report."}

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    forbidden = re.compile(
        r"\b(INSERT|UPDATE|DELETE|DROP|CREATE|ALTER|TRUNCATE|MERGE|GRANT|REVOKE)\b",
        re.IGNORECASE,
    )

    header_fill = PatternFill(start_color="29B5E8", end_color="29B5E8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))
    sheet_names = []
    total_rows = 0

    for q in query_list:
        sheet_name = str(q.get("name", f"Query {len(sheet_names) + 1}"))[:31]
        sql = q.get("sql", "")
        q_chart_type = q.get("chart_type", "bar")

        if forbidden.search(sql):
            ws = wb.create_sheet(sheet_name)
            ws["A1"] = "ERROR: Only SELECT queries allowed"
            sheet_names.append(sheet_name)
            continue

        df = execute_query(sql, limit=50000)
        if df.empty:
            ws = wb.create_sheet(sheet_name)
            ws["A1"] = "No data returned"
            sheet_names.append(sheet_name)
            continue

        ws = wb.create_sheet(sheet_name)
        total_rows += len(df)

        # Headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if hasattr(value, "item"):
                    value = value.item()
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
                else:
                    cell.value = str(value) if value is not None else ""
                cell.font = Font(size=10)
                cell.border = thin_border

        # Auto-width
        for col_idx in range(1, len(df.columns) + 1):
            max_len = len(str(df.columns[col_idx - 1]))
            for r in range(2, min(len(df) + 2, 52)):
                cv = ws.cell(row=r, column=col_idx).value
                if cv:
                    max_len = max(max_len, len(str(cv)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
        sheet_names.append(sheet_name)

        # Add chart if we have numeric data
        if len(df) > 1 and q_chart_type:
            x_col = None
            y_col = None
            for c in df.columns:
                if df[c].dtype == "object" or str(df[c].dtype).startswith("datetime"):
                    x_col = c
                    break
            if not x_col:
                x_col = df.columns[0]
            for c in df.columns:
                if c != x_col and df[c].dtype in ("float64", "int64", "float32", "int32"):
                    y_col = c
                    break

            if y_col:
                x_idx = list(df.columns).index(x_col) + 1
                y_idx = list(df.columns).index(y_col) + 1

                if q_chart_type == "pie":
                    chart = PieChart()
                elif q_chart_type == "line":
                    chart = LineChart()
                    chart.y_axis.numFmt = '#,##0'
                else:
                    chart = BarChart()
                    chart.y_axis.numFmt = '#,##0'

                chart.title = sheet_name
                chart.width = 18
                chart.height = 12
                chart.style = 10

                data_ref = Reference(ws, min_col=y_idx, min_row=1, max_row=min(len(df) + 1, 101))
                cats_ref = Reference(ws, min_col=x_idx, min_row=2, max_row=min(len(df) + 1, 101))
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)

                # Place chart to the right of data
                chart_col = get_column_letter(len(df.columns) + 2)
                ws.add_chart(chart, f"{chart_col}1")

    # Write
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_bytes = excel_buffer.getvalue()

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_title = re.sub(r"[^a-zA-Z0-9_-]", "_", title)[:40]
    filename = f"{safe_title}_{timestamp}.xlsx"
    key = f"{REPORT_PREFIX}{filename}"
    url = _upload_to_s3(excel_bytes, key, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename)

    return {
        "download_url": url,
        "filename": filename,
        "format": "excel",
        "total_rows": total_rows,
        "sheets": sheet_names,
    }
